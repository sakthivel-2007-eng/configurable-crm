"""Bulk member upload from Excel.

Dry run by default. A spreadsheet of people is the sort of thing an admin wants
to see the outcome of before it happens, and "created 40 members, 12 of them
wrong" has no undo in M1.

Rows are processed independently: one bad row reports an error and the rest
still import. Rejecting the whole file because row 37 has a typo makes people
edit spreadsheets in the dark.

Column headers are matched case-insensitively against a fixed set —
`email`, `full_name`, `template`, `manager_email`, `license`. These are
*product* columns describing a membership, not customer field definitions;
importing customer lead data is M7 and reads its columns from `lead_fields`.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from app.auth.passwords import PasswordHasherService
from app.errors import unprocessable
from app.models import Membership, PermissionTemplate, User
from app.schemas.member import BulkUploadReport, BulkUploadRow
from app.tenancy.session import ScopedSession

__all__ = ["MemberImportService"]

_MAX_ROWS = 2_000
_REQUIRED_COLUMNS = ("email", "full_name", "template")
_KNOWN_COLUMNS = (*_REQUIRED_COLUMNS, "manager_email", "license")


@dataclass(slots=True)
class _ParsedRow:
    row_number: int
    email: str | None
    full_name: str | None
    template: str | None
    manager_email: str | None
    grant_license: bool


class MemberImportService:
    """Parses an uploaded workbook and creates memberships from it."""

    def __init__(
        self,
        session: ScopedSession,
        *,
        hasher: PasswordHasherService,
    ) -> None:
        self._session = session
        self._hasher = hasher

    async def run(self, content: bytes, *, dry_run: bool = True) -> BulkUploadReport:
        rows = self._parse(content)
        templates = await self._templates_by_name()
        existing_emails = await self._existing_member_emails()

        results: list[BulkUploadRow] = []
        # Emails created earlier in this same file, so a duplicate row inside
        # one upload is reported rather than colliding at flush time.
        claimed: set[str] = set()

        for row in rows:
            outcome = await self._apply_row(
                row,
                templates=templates,
                existing_emails=existing_emails,
                claimed=claimed,
                dry_run=dry_run,
            )
            results.append(outcome)

        if dry_run:
            # Nothing was written, but ORM state may have been staged; drop it
            # so a subsequent commit in the same request cannot leak a preview.
            await self._session.rollback()

        return BulkUploadReport(
            dry_run=dry_run,
            total_rows=len(results),
            created=sum(1 for r in results if r.status == "created"),
            skipped=sum(1 for r in results if r.status == "skipped"),
            errored=sum(1 for r in results if r.status == "error"),
            rows=results,
        )

    # --- parsing -----------------------------------------------------------

    def _parse(self, content: bytes) -> list[_ParsedRow]:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises a wide variety on bad input
            raise unprocessable(
                "invalid_workbook",
                "The uploaded file could not be read as an Excel workbook",
            ) from exc

        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(values_only=True)

            try:
                header_row = next(rows)
            except StopIteration:
                raise unprocessable("empty_workbook", "The workbook has no rows") from None

            headers = self._map_headers(header_row)
            missing = [column for column in _REQUIRED_COLUMNS if column not in headers]
            if missing:
                raise unprocessable(
                    "missing_columns",
                    f"The workbook is missing required columns: {', '.join(missing)}",
                    missing_columns=missing,
                )

            parsed: list[_ParsedRow] = []
            for offset, raw in enumerate(rows, start=2):
                if offset - 1 > _MAX_ROWS:
                    raise unprocessable(
                        "too_many_rows",
                        f"A member upload is limited to {_MAX_ROWS} rows",
                    )
                if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
                    continue
                parsed.append(self._parse_row(offset, raw, headers))
            return parsed
        finally:
            workbook.close()

    @staticmethod
    def _map_headers(header_row: tuple[object, ...]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for index, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().casefold().replace(" ", "_")
            if key in _KNOWN_COLUMNS:
                mapped[key] = index
        return mapped

    @staticmethod
    def _parse_row(
        row_number: int,
        raw: tuple[object, ...],
        headers: dict[str, int],
    ) -> _ParsedRow:
        def cell(name: str) -> str | None:
            index = headers.get(name)
            if index is None or index >= len(raw):
                return None
            value = raw[index]
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        license_value = (cell("license") or "").casefold()
        return _ParsedRow(
            row_number=row_number,
            email=(cell("email") or "").casefold() or None,
            full_name=cell("full_name"),
            template=cell("template"),
            manager_email=(cell("manager_email") or "").casefold() or None,
            grant_license=license_value in ("yes", "true", "1", "y"),
        )

    # --- application -------------------------------------------------------

    async def _apply_row(
        self,
        row: _ParsedRow,
        *,
        templates: dict[str, PermissionTemplate],
        existing_emails: set[str],
        claimed: set[str],
        dry_run: bool,
    ) -> BulkUploadRow:
        def outcome(status: str, message: str | None = None) -> BulkUploadRow:
            return BulkUploadRow(
                row_number=row.row_number,
                email=row.email,
                full_name=row.full_name,
                template_name=row.template,
                manager_email=row.manager_email,
                status=status,
                message=message,
            )

        if not row.email:
            return outcome("error", "email is required")
        if not row.full_name:
            return outcome("error", "full_name is required")
        if not row.template:
            return outcome("error", "template is required")

        template = templates.get(row.template.casefold())
        if template is None:
            return outcome("error", f"No permission template named {row.template!r}")

        if row.email in existing_emails or row.email in claimed:
            return outcome("skipped", "Already a member of this workspace")

        manager_id: uuid.UUID | None = None
        if row.manager_email:
            manager_id = await self._membership_id_for_email(row.manager_email)
            if manager_id is None:
                return outcome("error", f"No member found for manager_email {row.manager_email!r}")

        claimed.add(row.email)

        if dry_run:
            return outcome("created", "Would be created")

        user = await self._get_or_create_user(email=row.email, full_name=row.full_name)
        membership = Membership(
            user_id=user.id,
            template_id=template.id,
            manager_id=manager_id,
            has_license=row.grant_license,
        )
        self._session.add(membership)
        await self._session.flush()
        return outcome("created")

    async def _get_or_create_user(self, *, email: str, full_name: str) -> User:
        """Find the user, or create one who cannot yet sign in.

        A bulk-created user gets an unguessable random password rather than a
        blank or shared one — they reach the product through a password reset,
        never through a credential that was in a spreadsheet.
        """
        result = await self._session.execute(select(User).where(User.email == email).limit(1))
        existing: User | None = result.scalar_one_or_none()
        if existing is not None:
            return existing

        user = User(
            email=email,
            full_name=full_name,
            password_hash=self._hasher.hash(uuid.uuid4().hex + uuid.uuid4().hex),
        )
        # A User is not tenant data — it deliberately does not go through
        # ScopedSession.add, which only accepts TenantModel.
        self._session.add_global(user)
        await self._session.flush()
        return user

    async def _templates_by_name(self) -> dict[str, PermissionTemplate]:
        rows = await self._session.execute(self._session.select(PermissionTemplate))
        return {template.name.casefold(): template for template in rows.scalars().all()}

    async def _existing_member_emails(self) -> set[str]:
        rows = await self._session.execute(
            self._session.select(Membership).join(User, Membership.user_id == User.id)
        )
        memberships = rows.scalars().all()
        if not memberships:
            return set()
        user_rows = await self._session.execute(
            select(User.email).where(User.id.in_([m.user_id for m in memberships]))
        )
        return {email.casefold() for email in user_rows.scalars().all()}

    async def _membership_id_for_email(self, email: str) -> uuid.UUID | None:
        rows = await self._session.execute(
            self._session.select(Membership)
            .join(User, Membership.user_id == User.id)
            .where(User.email == email)
            .limit(1)
        )
        membership = rows.scalar_one_or_none()
        return membership.id if membership else None
