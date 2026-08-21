"""Reading an uploaded spreadsheet (M7).

One reader for CSV and XLSX, shared by task bulk-upload, lead import and the
historical action import. Three flows, one parser: they disagree about what the
columns *mean*, never about how to get at them.

The reader is deliberately forgiving about shape and strict about size. A
customer's export from their previous CRM will have blank rows, stray trailing
columns, numbers Excel decided were dates, and a header row with inconsistent
capitalisation — none of which is a reason to refuse the file. What it will not
do is read an unbounded number of rows into memory.
"""

from __future__ import annotations

import csv
import dataclasses
import datetime as dt
import io
from collections.abc import Iterator, Sequence
from typing import Any

from openpyxl import load_workbook

from app.errors import unprocessable

__all__ = ["MAX_ROWS", "Sheet", "read_sheet"]

#: Rows accepted from one upload. Above this the operator should be splitting
#: the file, and below it a dry run stays fast enough to be interactive.
MAX_ROWS = 20_000

#: Header cells to read before giving up. A file with more columns than this is
#: not a lead list.
MAX_COLUMNS = 200


@dataclasses.dataclass(frozen=True, slots=True)
class Sheet:
    """A parsed upload: the header row, and the rows under it."""

    columns: tuple[str, ...]
    rows: tuple[dict[str, str], ...]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _clean(value: Any) -> str:
    """One cell, as the text a human would have typed.

    Excel hands back `datetime` for anything it decided was a date and `float`
    for anything numeric, so `7` arrives as `7.0` and would be written into a
    NUMBER field as the string "7.0". Normalising here means the field
    validators downstream see what the operator saw in the spreadsheet.
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        # Midnight almost always means "this was a date, not an instant".
        if value.time() == dt.time.min:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dedupe(headers: Sequence[str]) -> tuple[str, ...]:
    """Make the header row usable as a set of keys.

    Blank headers become positional names and duplicates get a suffix, because
    a mapping UI keyed on column name cannot address two columns called
    "Phone" — and refusing the file over it would be unhelpful when the second
    one is empty anyway.
    """
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, raw in enumerate(headers):
        name = raw.strip() or f"Column {index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        result.append(name)
    return tuple(result)


def _rows_from_csv(content: bytes) -> Iterator[list[str]]:
    # utf-8-sig: Excel's CSV export leads with a BOM, which would otherwise
    # become part of the first column's name and break every mapping.
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        # Windows-1252 is what a non-UTF-8 CSV from Excel almost always is.
        text = content.decode("cp1252", errors="replace")
    yield from csv.reader(io.StringIO(text))


def _rows_from_xlsx(content: bytes) -> Iterator[list[str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise unprocessable(
            "invalid_workbook", "That file could not be read as a spreadsheet"
        ) from exc

    sheet = workbook.active
    if sheet is None:  # pragma: no cover - openpyxl always gives one
        raise unprocessable("invalid_workbook", "That workbook has no sheets")

    for row in sheet.iter_rows(values_only=True):
        yield [_clean(cell) for cell in row[:MAX_COLUMNS]]
    workbook.close()


def read_sheet(content: bytes, *, filename: str) -> Sheet:
    """Parse an upload into a header row and dictionaries under it.

    Chooses the parser by extension rather than by sniffing: an operator who
    renamed a `.csv` to `.xlsx` has made a mistake worth telling them about.
    """
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        raw_rows = _rows_from_csv(content)
    elif lowered.endswith((".xlsx", ".xlsm")):
        raw_rows = _rows_from_xlsx(content)
    else:
        raise unprocessable(
            "unsupported_file_type",
            "Upload a .csv or .xlsx file",
            filename=filename,
        )

    iterator = iter(raw_rows)
    try:
        header = next(iterator)
    except StopIteration:
        raise unprocessable("empty_file", "That file has no rows") from None

    columns = _dedupe([_clean(cell) for cell in header])
    if not columns:
        raise unprocessable("empty_file", "That file has no header row")

    rows: list[dict[str, str]] = []
    for raw in iterator:
        cells = [_clean(cell) for cell in raw]
        # A row of nothing is Excel padding, not data. Skipping it silently is
        # right: reporting a hundred "row 4012 was blank" warnings would bury
        # the ones that matter.
        if not any(cells):
            continue
        if len(rows) >= MAX_ROWS:
            raise unprocessable(
                "too_many_rows",
                f"That file has more than {MAX_ROWS:,} rows. Split it and upload again.",
                limit=MAX_ROWS,
            )
        # `zip` truncates to the shorter side, which handles both a row with
        # trailing junk and a row that stops early.
        rows.append(dict(zip(columns, cells, strict=False)))

    return Sheet(columns=columns, rows=tuple(rows))
